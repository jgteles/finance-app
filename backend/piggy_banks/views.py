from decimal import Decimal, InvalidOperation
from datetime import datetime

from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from django.utils.text import slugify
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import PiggyBank, PiggyBankMovement
from .serializers import PiggyBankMovementSerializer, PiggyBankSerializer


class PiggyBankViewSet(viewsets.ModelViewSet):
    queryset = PiggyBank.objects.all()
    serializer_class = PiggyBankSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return PiggyBank.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def _parse_amount(self, request) -> Decimal | None:
        raw = request.data.get("amount", None)
        if raw is None:
            return None
        try:
            return Decimal(str(raw))
        except (InvalidOperation, TypeError, ValueError):
            return None

    @action(detail=True, methods=["post"])
    def deposit(self, request, pk=None):
        amount = self._parse_amount(request)
        if amount is None or amount <= 0:
            return Response(
                {"error": "amount must be a positive number"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        piggy = self.get_object()
        with transaction.atomic():
            piggy.balance = piggy.balance + amount
            piggy.save(update_fields=["balance", "updated_at"])
            PiggyBankMovement.objects.create(
                piggy_bank=piggy,
                movement_type=PiggyBankMovement.MovementType.DEPOSIT,
                amount=amount,
                balance_after=piggy.balance,
            )

        return Response(self.get_serializer(piggy).data)

    @action(detail=True, methods=["post"])
    def withdraw(self, request, pk=None):
        amount = self._parse_amount(request)
        if amount is None or amount <= 0:
            return Response(
                {"error": "amount must be a positive number"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        piggy = self.get_object()
        next_balance = piggy.balance - amount
        if next_balance < 0:
            return Response(
                {"error": "insufficient balance"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            piggy.balance = next_balance
            piggy.save(update_fields=["balance", "updated_at"])
            PiggyBankMovement.objects.create(
                piggy_bank=piggy,
                movement_type=PiggyBankMovement.MovementType.WITHDRAW,
                amount=amount,
                balance_after=piggy.balance,
            )

        return Response(self.get_serializer(piggy).data)

    @action(detail=True, methods=["get"])
    def movements(self, request, pk=None):
        piggy = self.get_object()
        queryset = piggy.movements.all()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = PiggyBankMovementSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = PiggyBankMovementSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"], url_path="movements-export-excel")
    def export_movements_excel(self, request, pk=None):
        piggy = self.get_object()
        movements = piggy.movements.order_by("created_at", "id").all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimentações"

        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = piggy.name
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill("solid", fgColor="2563EB")
        ws.row_dimensions[1].height = 28

        ws.append([])  # Row 2 spacing
        ws.append(["Data", "Tipo", "Valor"])

        header_fill = PatternFill("solid", fgColor="E2E8F0")
        header_font = Font(bold=True, color="0F172A")
        for col in ("A", "B", "C"):
            cell = ws[f"{col}3"]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        total_deposit = Decimal("0")
        total_withdraw = Decimal("0")

        for m in movements:
            created_at = m.created_at
            if timezone.is_aware(created_at):
                created_at = timezone.localtime(created_at).replace(tzinfo=None)
            tipo = "Aporte" if m.movement_type == PiggyBankMovement.MovementType.DEPOSIT else "Retirada"
            ws.append([created_at, tipo, float(m.amount)])

            if m.movement_type == PiggyBankMovement.MovementType.DEPOSIT:
                total_deposit += m.amount
            else:
                total_withdraw += m.amount

        ws.append([])
        ws.append(["Total aportado", "", float(total_deposit)])
        ws.append(["Total retirado", "", float(total_withdraw)])
        ws.append(["Saldo do cofrinho", "", float(piggy.balance)])

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 14

        data_end_row = 3 + len(movements)
        for row in ws.iter_rows(min_row=4, max_row=data_end_row, min_col=1, max_col=3):
            if isinstance(row[0].value, datetime):
                row[0].number_format = "dd/mm/yyyy hh:mm:ss"
            row[1].alignment = Alignment(horizontal="center")
            row[2].number_format = '"R$" #,##0.00'

        last_row = ws.max_row
        totals_start = last_row - 2
        for r in range(totals_start, last_row + 1):
            ws[f"A{r}"].font = Font(bold=True)
            ws[f"C{r}"].font = Font(bold=True)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        safe_name = slugify(piggy.name) or f"cofrinho-{piggy.id}"
        response["Content-Disposition"] = (
            f'attachment; filename="{safe_name}-movimentacoes.xlsx"'
        )
        wb.save(response)
        return response

    @action(
        detail=True,
        methods=["delete"],
        url_path=r"movements/(?P<movement_id>[^/.]+)",
    )
    def delete_movement(self, request, pk=None, movement_id=None):
        piggy = self.get_object()

        with transaction.atomic():
            piggy = PiggyBank.objects.select_for_update().get(pk=piggy.pk, user=request.user)

            movement = piggy.movements.filter(id=movement_id).first()
            if movement is None:
                return Response(
                    {"error": "movement not found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            ordered = list(piggy.movements.order_by("created_at", "id").all())
            net = Decimal("0")
            for m in ordered:
                if m.movement_type == PiggyBankMovement.MovementType.DEPOSIT:
                    net += m.amount
                else:
                    net -= m.amount

            baseline = piggy.balance - net

            movement.delete()

            remaining = list(piggy.movements.order_by("created_at", "id").all())
            running_balance = baseline
            for m in remaining:
                if m.movement_type == PiggyBankMovement.MovementType.DEPOSIT:
                    running_balance += m.amount
                else:
                    running_balance -= m.amount

                if running_balance < 0:
                    return Response(
                        {
                            "error": "deleting this movement would result in negative balance due to later withdrawals"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                m.balance_after = running_balance

            if remaining:
                PiggyBankMovement.objects.bulk_update(remaining, ["balance_after"])

            piggy.balance = running_balance
            piggy.save(update_fields=["balance", "updated_at"])

        return Response(status=status.HTTP_204_NO_CONTENT)
